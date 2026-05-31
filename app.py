import os
import base64
import json
import re
import tempfile
from datetime import datetime
from io import BytesIO
from PIL import Image as PILImage

from flask import Flask, request, jsonify, render_template, send_file
from dotenv import load_dotenv
import anthropic
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

load_dotenv()

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 50 * 1024 * 1024  # 50MB

client = anthropic.Anthropic(api_key=os.environ.get("ANTHROPIC_API_KEY"))

WORK_CATEGORIES = [
    "クロス張替え（全面）",
    "クロス張替え（部分）",
    "クッションフロア張替え",
    "フローリング補修・部分貼替え",
    "畳替え",
    "ハウスクリーニング（基本）",
    "エアコンクリーニング",
    "浴室・強化クリーニング",
    "照明器具交換・修繕",
    "給湯器・設備修繕",
    "鍵交換",
    "リペア補修（床・壁複数箇所）",
    "残置物・廃棄物処分",
    "消毒・防虫処理",
    "諸経費（パーキング・交通費等）",
    "その他（自由入力）",
]

DEFAULT_BURDEN = {
    "クロス張替え（全面）": "貸",
    "クロス張替え（部分）": "借",
    "クッションフロア張替え": "貸",
    "フローリング補修・部分貼替え": "両",
    "畳替え": "両",
    "ハウスクリーニング（基本）": "借",
    "エアコンクリーニング": "借",
    "浴室・強化クリーニング": "借",
    "照明器具交換・修繕": "貸",
    "給湯器・設備修繕": "貸",
    "鍵交換": "借",
    "リペア補修（床・壁複数箇所）": "借",
    "残置物・廃棄物処分": "借",
    "消毒・防虫処理": "借",
    "諸経費（パーキング・交通費等）": "貸",
    "その他（自由入力）": "借",
}


# 東京都ガイドラインのデフォルト負担（特約なし時）
TOKYO_GUIDELINE_BURDEN = {
    "クロス張替え（全面）":        {"burden": "貸", "reason": "経年劣化・自然消耗（東京都ガイドライン）"},
    "クロス張替え（部分）":        {"burden": "借", "reason": "故意・過失による損傷（東京都ガイドライン）"},
    "クッションフロア張替え":      {"burden": "借", "reason": "借主使用による損耗・6年償却（東京都ガイドライン）"},
    "フローリング補修・部分貼替え":{"burden": "両", "reason": "程度・原因による折半（東京都ガイドライン）"},
    "畳替え":                     {"burden": "貸", "reason": "通常使用による交換（東京都ガイドライン）"},
    "ハウスクリーニング（基本）":  {"burden": "貸", "reason": "通常清掃実施の場合は貸主負担（東京都ガイドライン）"},
    "エアコンクリーニング":        {"burden": "貸", "reason": "臭い等なければ貸主負担（東京都ガイドライン）"},
    "浴室・強化クリーニング":      {"burden": "借", "reason": "清掃怠慢による汚損（東京都ガイドライン）"},
    "照明器具交換・修繕":          {"burden": "貸", "reason": "設備の経年交換（東京都ガイドライン）"},
    "給湯器・設備修繕":            {"burden": "貸", "reason": "設備の経年劣化（東京都ガイドライン）"},
    "鍵交換":                     {"burden": "借", "reason": "紛失・破損による交換（東京都ガイドライン）"},
    "リペア補修（床・壁複数箇所）":{"burden": "借", "reason": "故意・過失による損傷（東京都ガイドライン）"},
    "残置物・廃棄物処分":          {"burden": "借", "reason": "借主残置（東京都ガイドライン）"},
    "消毒・防虫処理":              {"burden": "借", "reason": "借主使用による必要（東京都ガイドライン）"},
    "諸経費（パーキング・交通費等）":{"burden": "貸", "reason": "工事付帯費用（東京都ガイドライン）"},
    "その他（自由入力）":          {"burden": "借", "reason": "内容確認が必要"},
}


@app.route("/")
def index():
    return render_template("index.html", categories=WORK_CATEGORIES)


@app.route("/api/analyze-contract", methods=["POST"])
def analyze_contract():
    """賃貸借契約書PDFを解析して特約内容と負担区分を返す"""
    if "pdf" not in request.files:
        return jsonify({"error": "PDFファイルが見つかりません"}), 400
    pdf_file = request.files["pdf"]
    if pdf_file.filename == "":
        return jsonify({"error": "ファイルが選択されていません"}), 400
    try:
        pdf_bytes = pdf_file.read()
        pdf_b64 = base64.standard_b64encode(pdf_bytes).decode("utf-8")

        categories_str = "\n".join(f"- {c}" for c in WORK_CATEGORIES)
        guideline_str = "\n".join(
            f'- {k}: デフォルト{v["burden"]}（{v["reason"]}）'
            for k, v in TOKYO_GUIDELINE_BURDEN.items()
        )

        prompt = f"""この賃貸借契約書（特に「頭書(8)特約事項」「第16条 明渡し時の原状回復」）を詳しく読んでください。

以下の工事項目について、特約事項の記載を確認し、各項目の負担区分を判定してください。

【工事項目リスト】
{categories_str}

【判定ルール】
1. 特約事項に明記がある場合 → 特約の内容に従う（basis: "契約特約"）
2. 特約事項に記載がない場合 → 東京都ガイドラインに従う（basis: "東京都ガイドライン"）

【東京都ガイドライン デフォルト値】
{guideline_str}

【重要な判定ポイント】
- 「クリーニング費用を借主負担とする」特約があれば → ハウスクリーニング・エアコンクリーニング・浴室クリーニング等 → 借主負担
- 「使用状況にかかわらず」という記載があれば → ガイドライン原則を上書きして借主負担
- 「故意・過失による場合のみ」という記載 → ガイドラインに従う
- 禁煙特約がある場合 → タバコ関連は借主負担

JSON形式のみで回答してください（説明文不要）：
{{
  "tenant_name": "借主氏名（不明なら空文字）",
  "landlord_name": "貸主氏名（不明なら空文字）",
  "property_name": "物件名（不明なら空文字）",
  "special_clauses": ["特約の主要内容を簡潔に列挙"],
  "burden_assignments": {{
    "クロス張替え（全面）": {{"burden": "貸", "basis": "東京都ガイドライン", "reason": "経年劣化"}},
    "クロス張替え（部分）": {{"burden": "借", "basis": "東京都ガイドライン", "reason": "故意・過失"}},
    "クッションフロア張替え": {{"burden": "借", "basis": "東京都ガイドライン", "reason": "借主使用"}},
    "フローリング補修・部分貼替え": {{"burden": "両", "basis": "東京都ガイドライン", "reason": "程度による"}},
    "畳替え": {{"burden": "貸", "basis": "東京都ガイドライン", "reason": "通常交換"}},
    "ハウスクリーニング（基本）": {{"burden": "借 or 貸", "basis": "契約特約 or 東京都ガイドライン", "reason": "理由"}},
    "エアコンクリーニング": {{"burden": "借 or 貸", "basis": "契約特約 or 東京都ガイドライン", "reason": "理由"}},
    "浴室・強化クリーニング": {{"burden": "借", "basis": "東京都ガイドライン", "reason": "清掃怠慢"}},
    "照明器具交換・修繕": {{"burden": "貸", "basis": "東京都ガイドライン", "reason": "設備経年"}},
    "給湯器・設備修繕": {{"burden": "貸", "basis": "東京都ガイドライン", "reason": "設備経年"}},
    "鍵交換": {{"burden": "借", "basis": "東京都ガイドライン", "reason": "紛失・破損"}},
    "リペア補修（床・壁複数箇所）": {{"burden": "借", "basis": "東京都ガイドライン", "reason": "故意・過失"}},
    "残置物・廃棄物処分": {{"burden": "借", "basis": "東京都ガイドライン", "reason": "借主残置"}},
    "消毒・防虫処理": {{"burden": "借", "basis": "東京都ガイドライン", "reason": "借主使用"}},
    "諸経費（パーキング・交通費等）": {{"burden": "貸", "basis": "東京都ガイドライン", "reason": "工事付帯費"}},
    "その他（自由入力）": {{"burden": "借", "basis": "東京都ガイドライン", "reason": "要確認"}}
  }}
}}

burden値は必ず "貸"、"借"、"両" のいずれか。"""

        message = client.messages.create(
            model="claude-sonnet-4-5",
            max_tokens=4096,
            messages=[
                {
                    "role": "user",
                    "content": [
                        {
                            "type": "document",
                            "source": {
                                "type": "base64",
                                "media_type": "application/pdf",
                                "data": pdf_b64,
                            },
                        },
                        {"type": "text", "text": prompt},
                    ],
                }
            ],
        )

        response_text = message.content[0].text.strip()
        json_match = re.search(r"\{.*\}", response_text, re.DOTALL)
        if json_match:
            response_text = json_match.group()

        result = json.loads(response_text)
        return jsonify(result)

    except json.JSONDecodeError as e:
        return jsonify({"error": f"解析結果のパースエラー: {str(e)}"}), 500
    except anthropic.APIError as e:
        return jsonify({"error": f"Claude APIエラー: {str(e)}"}), 500
    except Exception as e:
        return jsonify({"error": f"契約書解析エラー: {str(e)}"}), 500


@app.route("/api/analyze", methods=["POST"])
def analyze():
    if "pdf" not in request.files:
        return jsonify({"error": "PDFファイルが見つかりません"}), 400

    pdf_file = request.files["pdf"]
    if pdf_file.filename == "":
        return jsonify({"error": "ファイルが選択されていません"}), 400

    try:
        pdf_bytes = pdf_file.read()
        pdf_b64 = base64.standard_b64encode(pdf_bytes).decode("utf-8")

        categories_str = "\n".join(f"- {c}" for c in WORK_CATEGORIES)

        prompt = f"""この見積書PDFを解析して、以下の情報をJSON形式で返してください。

抽出する情報：
1. 物件情報：
   - property_name: 物件名（不明な場合は空文字）
   - estimate_number: 見積番号（不明な場合は空文字）
   - estimate_date: 見積日（YYYY-MM-DD形式、不明な場合は空文字）
   - company_name: 業者名（不明な場合は空文字）
   - staff_name: 担当者名（不明な場合は空文字）

2. 工事明細（itemsリスト）：
   - category: 以下の工事分類から最も近いものを1つ選ぶ
{categories_str}
   - description: 仕様・備考（元の記載をそのまま）
   - quantity: 数量（数値）
   - unit: 単位（㎡、式、ヶ所、台など）
   - unit_price: 単価（税抜、数値）

レスポンスは必ず以下のJSON形式のみで返してください（説明文不要）：
{{
  "property_name": "",
  "estimate_number": "",
  "estimate_date": "",
  "company_name": "",
  "staff_name": "",
  "items": [
    {{
      "category": "",
      "description": "",
      "quantity": 0,
      "unit": "",
      "unit_price": 0
    }}
  ]
}}"""

        message = client.messages.create(
            model="claude-sonnet-4-5",
            max_tokens=4096,
            messages=[
                {
                    "role": "user",
                    "content": [
                        {
                            "type": "document",
                            "source": {
                                "type": "base64",
                                "media_type": "application/pdf",
                                "data": pdf_b64,
                            },
                        },
                        {
                            "type": "text",
                            "text": prompt,
                        },
                    ],
                }
            ],
        )

        response_text = message.content[0].text.strip()

        # JSON部分を抽出
        json_match = re.search(r"\{.*\}", response_text, re.DOTALL)
        if json_match:
            response_text = json_match.group()

        result = json.loads(response_text)

        # 負担区分のデフォルト値を付与
        for item in result.get("items", []):
            category = item.get("category", "その他（自由入力）")
            item["burden"] = DEFAULT_BURDEN.get(category, "借")

        return jsonify(result)

    except json.JSONDecodeError as e:
        return jsonify({"error": f"APIレスポンスの解析に失敗しました: {str(e)}"}), 500
    except anthropic.APIError as e:
        return jsonify({"error": f"Claude APIエラー: {str(e)}"}), 500
    except Exception as e:
        return jsonify({"error": f"エラーが発生しました: {str(e)}"}), 500


@app.route("/api/export", methods=["POST"])
def export():
    try:
        # multipart（写真あり）とJSON両対応
        if request.content_type and request.content_type.startswith("multipart/form-data"):
            data = json.loads(request.form.get("data", "{}"))
            photos = {}  # {proposal_index: bytes}
            for key in request.files:
                if key.startswith("photo_"):
                    idx = int(key.split("_")[1])
                    photos[idx] = request.files[key].read()
        else:
            data = request.get_json()
            photos = {}

        if not data:
            return jsonify({"error": "データが見つかりません"}), 400

        wb = create_excel(data, photos=photos)

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        property_name = data.get("property_name", "物件名不明")
        date_str = datetime.now().strftime("%Y%m%d")
        filename = f"原状回復精算_{property_name}_{date_str}.xlsx"

        # output/フォルダにも保存
        output_dir = os.path.join(os.path.dirname(__file__), "output")
        os.makedirs(output_dir, exist_ok=True)
        save_path = os.path.join(output_dir, filename)
        with open(save_path, "wb") as f:
            output.seek(0)
            f.write(output.read())
        output.seek(0)

        encoded_filename = filename.encode("utf-8").decode("latin-1", errors="replace")

        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=filename,
        )

    except Exception as e:
        return jsonify({"error": f"Excel出力エラー: {str(e)}"}), 500


def make_border():
    side = Side(style="thin", color="000000")
    return Border(left=side, right=side, top=side, bottom=side)


def apply_header_style(cell, bg_color="1F3864"):
    cell.font = Font(name="Yu Gothic UI", bold=True, color="FFFFFF", size=10)
    cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
    cell.border = make_border()


def apply_data_style(cell, bg_color=None, bold=False, align="left", number_format=None):
    cell.font = Font(name="Yu Gothic UI", bold=bold, size=10)
    if bg_color:
        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border = make_border()
    if number_format:
        cell.number_format = number_format


def fill_row_borders(ws, row, start_col, end_col, bg_color=None):
    """指定行の全セルに罫線と背景色を設定"""
    for col in range(start_col, end_col + 1):
        c = ws.cell(row=row, column=col)
        c.border = make_border()
        if bg_color:
            c.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")


def create_excel(data, photos=None):
    wb = openpyxl.Workbook()
    if photos is None:
        photos = {}

    property_name    = data.get("property_name", "")
    estimate_number  = data.get("estimate_number", "")
    estimate_date    = data.get("estimate_date", "")
    company_name     = data.get("company_name", "")
    staff_name       = data.get("staff_name", "")
    tenant_name      = data.get("tenant_name", "")
    landlord_name    = data.get("landlord_name", "")
    property_address = data.get("property_address", "")
    deposit          = float(data.get("deposit", 0) or 0)
    landlord_rate    = float(data.get("landlord_rate", 1.15) or 1.15)
    tenant_rate      = float(data.get("tenant_rate", 1.20) or 1.20)
    tax_rate         = float(data.get("tax_rate", 0.10) or 0.10)
    items            = data.get("items", [])
    proposals        = data.get("proposals", [])

    # ① 負担割合表
    ws1 = wb.active
    ws1.title = "負担割合表"
    _build_estimate_sheet(ws1, data, items, landlord_rate, tenant_rate, tax_rate,
                          property_name, estimate_number, estimate_date, company_name, staff_name,
                          tenant_name, property_address)

    # ② 貸主請求書
    ws2 = wb.create_sheet("貸主請求書")
    _build_invoice_sheet(ws2, data, items, landlord_rate, tenant_rate, tax_rate, mode="landlord")

    # ③ 借主請求書
    ws3 = wb.create_sheet("借主請求書")
    _build_invoice_sheet(ws3, data, items, landlord_rate, tenant_rate, tax_rate, mode="tenant")

    # ④ リノベーション提案書（提案がある場合）
    if proposals:
        ws4 = wb.create_sheet("リノベーション提案")
        _build_proposal_sheet(ws4, data, proposals, photos)

    return wb


COMPANY_INFO = "恵比寿不動産（株式会社ライフアドバンス）　賃貸管理事業部　〒150-0011 東京都渋谷区東3丁目25-11 TOKYU REIT恵比寿ビル4F　TEL: 03-6421-0544"


def _build_estimate_sheet(ws, data, items, landlord_rate, tenant_rate, tax_rate,
                           property_name, estimate_number, estimate_date, company_name, staff_name,
                           tenant_name="", property_address=""):
    # 12列構成: A=No B=名称 C=施工箇所 D=数量 E=単位 F=単価 G=金額
    #           H=借主割合 I=借主金額 J=貸主割合 K=貸主金額 L=備考
    col_w = {"A":5,"B":26,"C":20,"D":7,"E":6,"F":11,"G":13,"H":7,"I":13,"J":7,"K":13,"L":28}
    for col, w in col_w.items():
        ws.column_dimensions[col].width = w

    issue_date = datetime.now().strftime("%Y年%m月%d日")

    # ── Row1: 発行日 ──────────────────────────────
    ws["A1"] = f"発行日：{issue_date}"
    ws["A1"].font = Font(name="Yu Gothic UI", size=9, color="555555")
    ws.row_dimensions[1].height = 14

    # ── Row2: タイトル（中央）＋ 会社名（右） ────────────
    ws.merge_cells("C2:G2")
    c = ws["C2"]; c.value = "お見積書"
    c.font = Font(name="Yu Gothic UI", bold=True, size=20, color="1F3864")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("H2:L2")
    c = ws["H2"]; c.value = "株式会社 ライフアドバンス"
    c.font = Font(name="Yu Gothic UI", bold=True, size=11, color="1F3864")
    c.alignment = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[2].height = 34

    # ── Row3: 借主名（左）＋ 会社住所（右） ──────────────
    ws.merge_cells("A3:E3")
    c = ws["A3"]; c.value = f"{tenant_name}　様" if tenant_name else "　様"
    c.font = Font(name="Yu Gothic UI", bold=True, size=14)
    c.alignment = Alignment(horizontal="left", vertical="bottom")
    ws.merge_cells("H3:L3")
    c = ws["H3"]; c.value = "所在地：東京都渋谷区東3-25-11 TOKYU REIT恵比寿ビル4階"
    c.font = Font(name="Yu Gothic UI", size=8, color="666666")
    c.alignment = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[3].height = 24

    # ── Row4: お見積もり金額（貸主負担税込） ─────────────
    landlord_ex_preview = 0
    for item in items:
        qty = float(item.get("quantity", 0) or 0)
        price = float(item.get("unit_price", 0) or 0)
        vendor = int(qty * price)
        burden = item.get("burden", "借")
        if burden == "貸":
            landlord_ex_preview += int(vendor * landlord_rate)
        elif burden == "両":
            landlord_ex_preview += int((vendor / 2) * landlord_rate)
    landlord_tax_preview = int(landlord_ex_preview * (1 + tax_rate))

    ws.merge_cells("A4:B4")
    c = ws["A4"]; c.value = "お見積もり金額(税込)"
    c.font = Font(name="Yu Gothic UI", size=9, bold=True)
    c.fill = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = make_border()
    ws.merge_cells("C4:F4")
    c = ws["C4"]; c.value = landlord_tax_preview
    c.number_format = '#,##0"円"'
    c.font = Font(name="Yu Gothic UI", bold=True, size=14, color="1F3864")
    c.fill = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = make_border()
    ws.row_dimensions[4].height = 26

    # ── Row5-7: 工事情報 ──────────────────────────
    prop_rows = [("工事件名", "原状回復工事"), ("物件名", property_name), ("物件住所", property_address)]
    r = 5
    for label, value in prop_rows:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        c = ws.cell(row=r, column=1, value=label)
        c.font = Font(name="Yu Gothic UI", size=9, bold=True)
        c.fill = PatternFill(start_color="E8EEF8", end_color="E8EEF8", fill_type="solid")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = make_border()
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=8)
        c = ws.cell(row=r, column=3, value=value)
        c.font = Font(name="Yu Gothic UI", size=9)
        c.border = make_border()
        ws.row_dimensions[r].height = 16
        r += 1

    r += 1  # 空行
    ws.row_dimensions[r - 1].height = 6

    # ── テーブルヘッダー（2行） ───────────────────────
    # Row1: 単一セルは2行スパン、借主/貸主は2列スパン
    row_h1 = r
    for col, val in [(1,"No."),(2,"名称"),(3,"施工箇所"),(4,"数量"),(5,"単位"),(6,"単価"),(7,"金額"),(12,"備考")]:
        ws.merge_cells(start_row=row_h1, start_column=col, end_row=row_h1+1, end_column=col)
        c = ws.cell(row=row_h1, column=col, value=val)
        apply_header_style(c, "1F3864")
    ws.merge_cells(start_row=row_h1, start_column=8, end_row=row_h1, end_column=9)
    c = ws.cell(row=row_h1, column=8, value="借主様負担額")
    apply_header_style(c, "8B2500")
    ws.merge_cells(start_row=row_h1, start_column=10, end_row=row_h1, end_column=11)
    c = ws.cell(row=row_h1, column=10, value="貸主様負担額")
    apply_header_style(c, "1F3864")
    ws.row_dimensions[row_h1].height = 18

    # Row2: 割合/金額
    row_h2 = row_h1 + 1
    for col, val, bg in [(8,"割合","8B2500"),(9,"金額","8B2500"),(10,"割合","1F3864"),(11,"金額","1F3864")]:
        c = ws.cell(row=row_h2, column=col, value=val)
        apply_header_style(c, bg)
    ws.row_dimensions[row_h2].height = 16
    r = row_h2 + 1

    # ── データ行（20行固定） ──────────────────────────
    vendor_total = 0
    tenant_total_ex = 0
    landlord_total_ex = 0

    for i in range(1, 21):
        if i <= len(items):
            item = items[i - 1]
            qty = float(item.get("quantity", 0) or 0)
            price = float(item.get("unit_price", 0) or 0)
            vendor = int(qty * price)
            burden = item.get("burden", "借")
            if burden == "貸":
                t_ratio, l_ratio = "0%", "100%"
                t_amt = 0
                l_amt = int(vendor * landlord_rate)
            elif burden == "借":
                t_ratio, l_ratio = "100%", "0%"
                t_amt = int(vendor * tenant_rate)
                l_amt = 0
            elif burden == "両":
                t_ratio, l_ratio = "50%", "50%"
                t_amt = int((vendor / 2) * tenant_rate)
                l_amt = int((vendor / 2) * landlord_rate)
            else:
                t_ratio, l_ratio = "", ""
                t_amt, l_amt = 0, 0
            vendor_total += vendor
            tenant_total_ex += t_amt
            landlord_total_ex += l_amt
            bg = "FCE4D6" if burden == "借" else ("D9E1F2" if burden == "貸" else "E2EFDA")
            row_vals = [i, item.get("category",""), item.get("description",""),
                        qty, item.get("unit",""), price, vendor,
                        t_ratio, t_amt, l_ratio, l_amt, ""]
        else:
            bg = None
            row_vals = [i,"","","","","","","","","","",""]

        for col, val in enumerate(row_vals, 1):
            c = ws.cell(row=r, column=col, value=val if (i <= len(items) or col == 1) else None)
            c.font = Font(name="Yu Gothic UI", size=10)
            if bg:
                c.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
            c.border = make_border()
            if col in (6, 7, 9, 11) and i <= len(items):
                c.number_format = "#,##0"
            c.alignment = Alignment(
                horizontal="center" if col in (1, 4, 5, 8, 10) else ("right" if col in (6,7,9,11) else "left"),
                vertical="center"
            )
        ws.row_dimensions[r].height = 18
        r += 1

    # ── 小計・消費税・合計 ────────────────────────────
    tax_pct = int(tax_rate * 100)
    totals = [
        ("小計", vendor_total, tenant_total_ex, landlord_total_ex),
        (f"消費税　{tax_pct}%", int(vendor_total*tax_rate), int(tenant_total_ex*tax_rate), int(landlord_total_ex*tax_rate)),
        ("合計", int(vendor_total*(1+tax_rate)), int(tenant_total_ex*(1+tax_rate)), int(landlord_total_ex*(1+tax_rate))),
    ]
    for label, v_val, t_val, l_val in totals:
        fill_row_borders(ws, r, 1, 12, "FFF2CC")
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        c = ws.cell(row=r, column=1, value=label)
        c.font = Font(name="Yu Gothic UI", bold=True, size=10)
        c.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        c.alignment = Alignment(horizontal="right", vertical="center")
        c.border = make_border()
        for col, val in [(7,v_val),(9,t_val),(11,l_val)]:
            c = ws.cell(row=r, column=col, value=val)
            c.number_format = '#,##0"円"'
            apply_data_style(c, bg_color="FFF2CC", bold=True, align="right")
        ws.row_dimensions[r].height = 18
        r += 1

    # ── 弊社利益（内部確認用） ──────────────────────────
    profit = (landlord_total_ex + tenant_total_ex) - vendor_total
    profit_rate = (profit / vendor_total * 100) if vendor_total > 0 else 0
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=12)
    c = ws.cell(row=r, column=1, value=f"【内部】弊社利益（税抜）：¥{profit:,}　利益率：{profit_rate:.1f}%")
    apply_data_style(c, bg_color="FFE699", bold=True, align="center")
    ws.row_dimensions[r].height = 18
    r += 2

    # ── 備考 ──────────────────────────────────────
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=12)
    c = ws.cell(row=r, column=1, value="【備考】")
    c.font = Font(name="Yu Gothic UI", bold=True, size=10)
    c.border = make_border()
    ws.row_dimensions[r].height = 16
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r+2, end_column=12)
    c = ws.cell(row=r, column=1, value="")
    c.border = make_border()
    ws.row_dimensions[r].height = 20

    ws.freeze_panes = ws.cell(row=row_h2 + 1, column=1)


REPRESENTATIVE = "橿頭 知宏"


def _build_invoice_sheet(ws, data, items, landlord_rate, tenant_rate, tax_rate, mode="tenant"):
    """mode: 'tenant'=借主請求書, 'landlord'=貸主請求書"""
    tenant_name  = data.get("tenant_name", "")
    landlord_name = data.get("landlord_name", "")
    property_name = data.get("property_name", "")
    property_address = data.get("property_address", "")
    company_name = data.get("company_name", "")
    staff_name   = data.get("staff_name", "")
    addressee    = tenant_name if mode == "tenant" else landlord_name

    # 列幅 (12列: A-L)
    col_w = {"A":5,"B":26,"C":20,"D":7,"E":6,"F":11,"G":13,"H":7,"I":13,"J":7,"K":13,"L":28}
    for col, w in col_w.items():
        ws.column_dimensions[col].width = w

    issue_date = datetime.now().strftime("%Y年%m月%d日")

    # ── Row1: 発行日 ──────────────────────────────
    ws["A1"] = f"発行日：{issue_date}"
    ws["A1"].font = Font(name="Yu Gothic UI", size=9, color="555555")
    ws.row_dimensions[1].height = 14

    # ── Row2: タイトル（中央）────────────────────────
    ws.merge_cells("C2:G2")
    c = ws["C2"]; c.value = "ご請求書"
    c.font = Font(name="Yu Gothic UI", bold=True, size=20, color="1F3864")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 34

    # ── Row3: 宛名（左）──────────────────────────────
    ws.merge_cells("A3:E3")
    c = ws["A3"]; c.value = f"{addressee}　様" if addressee else "　　　　　　　　様"
    c.font = Font(name="Yu Gothic UI", bold=True, size=14)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[3].height = 26

    # 請求金額を事前計算
    tenant_ex = 0; landlord_ex = 0
    for item in items:
        qty   = float(item.get("quantity", 0) or 0)
        price = float(item.get("unit_price", 0) or 0)
        vendor = int(qty * price)
        burden = item.get("burden", "借")
        if burden == "貸":
            landlord_ex += int(vendor * landlord_rate)
        elif burden == "借":
            tenant_ex += int(vendor * tenant_rate)
        elif burden == "両":
            landlord_ex += int((vendor / 2) * landlord_rate)
            tenant_ex   += int((vendor / 2) * tenant_rate)
    req_ex  = tenant_ex if mode == "tenant" else landlord_ex
    req_tax = int(req_ex * (1 + tax_rate))  # 工事費総額（税込）

    # 借主請求書：敷金精算後の実請求額を算出
    deposit_val    = float(data.get("deposit", 0) or 0)
    effective_dep  = float(data.get("effective_deposit", deposit_val) or deposit_val)
    amort_months   = float(data.get("amortization_months", 0) or 0)
    amort_amount   = float(data.get("amortization_amount", 0) or 0)
    holder         = data.get("deposit_holder", "landlord")
    holder_label   = "当社（ライフアドバンス）" if holder == "company" else "貸主様"

    if mode == "tenant" and deposit_val > 0:
        # 差額 = 工事費 - 有効敷金
        dep_balance = int(req_tax) - int(effective_dep)
        is_refund   = dep_balance <= 0           # 敷金が余る → 返金
        net_amount  = abs(dep_balance)           # 実請求額 or 返金額
        box_label   = "返金予定額（税込）" if is_refund else "ご請求金額（敷金精算後・税込）"
        box_color   = "E8F5E9" if is_refund else "FFEBEE"
        box_text_color = "1B5E20" if is_refund else "B71C1C"
    else:
        dep_balance   = 0
        is_refund     = False
        net_amount    = req_tax
        box_label     = "ご請求金額（税込）"
        box_color     = "EEEEEE"
        box_text_color = "1F3864"

    # ── Row4: ご請求金額ボックス（敷金精算後の実額）───────
    ws.merge_cells("A4:B4")
    c = ws["A4"]; c.value = box_label
    c.font = Font(name="Yu Gothic UI", size=9, bold=True)
    c.fill = PatternFill(start_color=box_color, end_color=box_color, fill_type="solid")
    c.alignment = Alignment(horizontal="center", vertical="center"); c.border = make_border()
    ws.merge_cells("C4:G4")
    c = ws["C4"]; c.value = net_amount; c.number_format = '#,##0"円"'
    c.font = Font(name="Yu Gothic UI", bold=True, size=16, color=box_text_color)
    c.fill = PatternFill(start_color=box_color, end_color=box_color, fill_type="solid")
    c.alignment = Alignment(horizontal="center", vertical="center"); c.border = make_border()
    ws.row_dimensions[4].height = 30

    # ── 右側会社情報（Row2〜7, H〜L列）──────────────────
    co_rows = [
        (2, "株式会社 ライフアドバンス",         True,  12),
        (3, "東京都渋谷区東3-25-11",              False,  9),
        (4, "TOKYU REIT恵比寿ビル4階",            False,  9),
        (5, "TEL03-6421-0544",                   False,  9),
        (6, f"施工会社：{company_name}",          False,  9),
        (7, f"施工担当：{staff_name}",            False,  9),
    ]
    for rn, text, bold, size in co_rows:
        ws.merge_cells(start_row=rn, start_column=9, end_row=rn, end_column=12)
        c = ws.cell(row=rn, column=9, value=text)
        c.font = Font(name="Yu Gothic UI", bold=bold, size=size,
                      color="1F3864" if bold else "444444")
        c.alignment = Alignment(horizontal="right", vertical="center")

    # ── Row5〜7: 工事情報（左側）──────────────────────
    prop_rows = [
        ("工事件名", "原状回復工事"),
        ("物件名",   property_name),
        ("物件住所", property_address),
    ]
    for idx, (label, value) in enumerate(prop_rows):
        rn = 5 + idx
        ws.merge_cells(start_row=rn, start_column=1, end_row=rn, end_column=2)
        c = ws.cell(row=rn, column=1, value=label)
        c.font = Font(name="Yu Gothic UI", size=9, bold=True)
        c.fill = PatternFill(start_color="E8EEF8", end_color="E8EEF8", fill_type="solid")
        c.alignment = Alignment(horizontal="center", vertical="center"); c.border = make_border()
        ws.merge_cells(start_row=rn, start_column=3, end_row=rn, end_column=8)
        c = ws.cell(row=rn, column=3, value=value)
        c.font = Font(name="Yu Gothic UI", size=9); c.border = make_border()
        ws.row_dimensions[rn].height = 16

    # ── Row8: スペーサー ─────────────────────────────
    ws.row_dimensions[8].height = 6

    # ── Rows9〜10: テーブルヘッダー（2行） ───────────────
    row_h1 = 9
    for col, val in [(1,"No."),(2,"名称"),(3,"施工箇所"),(4,"数量"),(5,"単位"),(6,"単価"),(7,"金額"),(12,"備考")]:
        ws.merge_cells(start_row=row_h1, start_column=col, end_row=row_h1+1, end_column=col)
        c = ws.cell(row=row_h1, column=col, value=val); apply_header_style(c, "1F3864")
    ws.merge_cells(start_row=row_h1, start_column=8, end_row=row_h1, end_column=9)
    c = ws.cell(row=row_h1, column=8, value="借主様負担額"); apply_header_style(c, "8B2500")
    ws.merge_cells(start_row=row_h1, start_column=10, end_row=row_h1, end_column=11)
    c = ws.cell(row=row_h1, column=10, value="貸主様負担額"); apply_header_style(c, "1F3864")
    ws.row_dimensions[row_h1].height = 20

    row_h2 = 10
    for col, val, bg in [(8,"割合","8B2500"),(9,"金額","8B2500"),(10,"割合","1F3864"),(11,"金額","1F3864")]:
        c = ws.cell(row=row_h2, column=col, value=val); apply_header_style(c, bg)
    ws.row_dimensions[row_h2].height = 16

    # ── Rows11〜30: データ行（20行固定）──────────────────
    vendor_total = 0; tenant_total_ex = 0; landlord_total_ex = 0

    for i in range(1, 21):
        r = 10 + i  # row 11〜30
        if i <= len(items):
            item   = items[i-1]
            qty    = float(item.get("quantity", 0) or 0)
            price  = float(item.get("unit_price", 0) or 0)
            vendor = int(qty * price)
            burden = item.get("burden", "借")
            if burden == "貸":
                t_ratio, l_ratio = "0%", "100%"
                t_amt = 0; l_amt = int(vendor * landlord_rate)
            elif burden == "借":
                t_ratio, l_ratio = "100%", "0%"
                t_amt = int(vendor * tenant_rate); l_amt = 0
            elif burden == "両":
                t_ratio, l_ratio = "50%", "50%"
                t_amt = int((vendor/2)*tenant_rate); l_amt = int((vendor/2)*landlord_rate)
            else:
                t_ratio, l_ratio = "0%", "0%"; t_amt = l_amt = 0
            vendor_total += vendor; tenant_total_ex += t_amt; landlord_total_ex += l_amt
            bg = "FCE4D6" if burden=="借" else ("D9E1F2" if burden=="貸" else "E2EFDA")
            row_vals = [i, item.get("category",""), item.get("description",""),
                        qty, item.get("unit",""), price, vendor, t_ratio, t_amt, l_ratio, l_amt, ""]
        else:
            bg = None
            row_vals = [i, "", "", "", "", "", "", "", 0, "", 0, ""]

        for col, val in enumerate(row_vals, 1):
            c = ws.cell(row=r, column=col)
            c.value = i if col == 1 else (val if val != "" else None)
            c.font = Font(name="Yu Gothic UI", size=10)
            if bg: c.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
            c.border = make_border()
            if col in (6,7,9,11): c.number_format = '#,##0"円"'
            c.alignment = Alignment(
                horizontal="center" if col in (1,4,5,8,10) else ("right" if col in (6,7,9,11) else "left"),
                vertical="center")
        ws.row_dimensions[r].height = 18

    # ── Rows31〜33: 小計・消費税・合計 ──────────────────
    tax_pct = int(tax_rate * 100)
    r = 31
    for label, v_val, t_val, l_val in [
        ("小計",               vendor_total,                   tenant_total_ex,                     landlord_total_ex),
        (f"消費税　{tax_pct}%", int(vendor_total*tax_rate),     int(tenant_total_ex*tax_rate),       int(landlord_total_ex*tax_rate)),
        ("合計",               int(vendor_total*(1+tax_rate)), int(tenant_total_ex*(1+tax_rate)),   int(landlord_total_ex*(1+tax_rate))),
    ]:
        fill_row_borders(ws, r, 1, 12, "FFF2CC")
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        c = ws.cell(row=r, column=1, value=label)
        c.font = Font(name="Yu Gothic UI", bold=True, size=10)
        c.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        c.alignment = Alignment(horizontal="right", vertical="center"); c.border = make_border()
        for col, val in [(7,v_val),(9,t_val),(11,l_val)]:
            c = ws.cell(row=r, column=col, value=val); c.number_format = '#,##0"円"'
            apply_data_style(c, bg_color="FFF2CC", bold=True, align="right")
        ws.row_dimensions[r].height = 18; r += 1

    # ── 敷金精算明細（借主請求書のみ）───────────────────
    if mode == "tenant" and deposit_val > 0:
        r34 = 34
        # セクションタイトル
        fill_row_borders(ws, r34, 1, 12, "E3F2FD")
        ws.merge_cells(start_row=r34, start_column=1, end_row=r34, end_column=12)
        c = ws.cell(row=r34, column=1, value="【敷金精算明細】")
        c.font = Font(name="Yu Gothic UI", bold=True, size=10, color="1F3864")
        c.fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
        c.border = make_border()
        ws.row_dimensions[r34].height = 18

        def dep_row(ws, rn, label, val, bold=False, bg="F8FFFE", color="000000", prefix=""):
            fill_row_borders(ws, rn, 1, 12, bg)
            ws.merge_cells(start_row=rn, start_column=1, end_row=rn, end_column=8)
            c = ws.cell(row=rn, column=1, value=label)
            c.font = Font(name="Yu Gothic UI", bold=bold, size=10, color=color)
            c.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
            c.alignment = Alignment(horizontal="right", vertical="center")
            c.border = make_border()
            ws.merge_cells(start_row=rn, start_column=9, end_row=rn, end_column=12)
            c2 = ws.cell(row=rn, column=9, value=val)
            c2.number_format = f'"{prefix}"#,##0"円"' if prefix else '#,##0"円"'
            c2.font = Font(name="Yu Gothic UI", bold=bold, size=11 if bold else 10, color=color)
            c2.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
            c2.alignment = Alignment(horizontal="right", vertical="center")
            c2.border = make_border()
            ws.row_dimensions[rn].height = 18

        rn = 35
        # 工事費（税込）
        dep_row(ws, rn, "原状回復工事費合計（借主負担・税込）", int(req_tax), bg="EEF5FF", bold=True)
        rn += 1
        # 敷金控除
        dep_row(ws, rn, f"△ 敷金（{holder_label}お預かり分）", int(deposit_val), bg="EEF5FF", prefix="▲ ")
        rn += 1
        if amort_amount > 0:
            dep_row(ws, rn, f"　　うち償却（{amort_months}ヶ月分・敷金より控除）", int(amort_amount), bg="F5F5FF", prefix="  ")
            rn += 1
        # 差引
        dep_row(ws, rn, "敷金充当後　差引残額", int(abs(dep_balance)), bg="DDEEFF", bold=True)
        rn += 1

        # 結果：返金 or 追加請求
        if is_refund:
            label  = f"返金額　（{holder_label}より借主様へご返金）"
            bg_bal = "E8F5E9"; col_bal = "1B5E20"
        else:
            label  = f"追加ご請求額　（上記振込先へお振込ください）"
            bg_bal = "FFEBEE"; col_bal = "B71C1C"

        fill_row_borders(ws, rn, 1, 12, bg_bal)
        ws.merge_cells(start_row=rn, start_column=1, end_row=rn, end_column=8)
        c = ws.cell(row=rn, column=1, value=label)
        c.font = Font(name="Yu Gothic UI", bold=True, size=11, color=col_bal)
        c.fill = PatternFill(start_color=bg_bal, end_color=bg_bal, fill_type="solid")
        c.alignment = Alignment(horizontal="right", vertical="center"); c.border = make_border()
        ws.merge_cells(start_row=rn, start_column=9, end_row=rn, end_column=12)
        c2 = ws.cell(row=rn, column=9, value=net_amount)
        c2.number_format = '#,##0"円"'
        c2.font = Font(name="Yu Gothic UI", bold=True, size=13, color=col_bal)
        c2.fill = PatternFill(start_color=bg_bal, end_color=bg_bal, fill_type="solid")
        c2.alignment = Alignment(horizontal="right", vertical="center"); c2.border = make_border()
        ws.row_dimensions[rn].height = 22
        rn += 2

        # 振込先セクション
        furikomi_start = rn
    else:
        furikomi_start = 34

    # ── 振込先・返金口座 ──────────────────────────────
    if mode == "tenant" and deposit_val > 0 and is_refund:
        furikomi_title = "【返金口座（借主様ご指定口座）】"
        furikomi_note  = "上記口座へ　　　年　　月　　日までにご返金いたします。"
    else:
        furikomi_title = "【振込先】"
        furikomi_note  = "上記口座へ　　　年　　月　　日までにお振込お願いいたします。"

    ws.merge_cells(start_row=furikomi_start, start_column=1, end_row=furikomi_start, end_column=12)
    c = ws.cell(row=furikomi_start, column=1, value=furikomi_title)
    c.font = Font(name="Yu Gothic UI", bold=True, size=10)
    c.border = make_border()
    ws.row_dimensions[furikomi_start].height = 18

    for rn in [furikomi_start+1, furikomi_start+2]:
        ws.merge_cells(start_row=rn, start_column=1, end_row=rn, end_column=12)
        ws.cell(row=rn, column=1).border = make_border()
        ws.row_dimensions[rn].height = 18

    ws.merge_cells(start_row=furikomi_start+3, start_column=1, end_row=furikomi_start+3, end_column=12)
    c = ws.cell(row=furikomi_start+3, column=1, value=furikomi_note)
    c.font = Font(name="Yu Gothic UI", size=10)
    c.border = make_border()
    ws.row_dimensions[furikomi_start+3].height = 22

    ws.freeze_panes = ws.cell(row=11, column=1)


def _build_proposal_sheet(ws, data, proposals, photos=None):
    """リノベーション提案書シートを生成"""
    if photos is None:
        photos = {}

    property_name    = data.get("property_name", "")
    property_address = data.get("property_address", "")
    staff_name       = data.get("staff_name", "")
    landlord_name    = data.get("landlord_name", "")
    issue_date       = datetime.now().strftime("%Y年%m月%d日")

    # 列幅設定
    col_w = {"A":5,"B":28,"C":28,"D":22,"E":14,"F":22}
    for col, w in col_w.items():
        ws.column_dimensions[col].width = w

    # ── Row1: 発行日
    ws["A1"] = f"発行日：{issue_date}"
    ws["A1"].font = Font(name="Yu Gothic UI", size=9, color="555555")
    ws.row_dimensions[1].height = 14

    # ── Row2: タイトル
    ws.merge_cells("B2:E2")
    c = ws["B2"]; c.value = "リノベーション提案書"
    c.font = Font(name="Yu Gothic UI", bold=True, size=20, color="2E7D32")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 34

    # ── Row3: 宛名（左）+ 会社名（右）
    ws.merge_cells("A3:C3")
    c = ws["A3"]; c.value = f"{landlord_name}　様" if landlord_name else "　　　　　様"
    c.font = Font(name="Yu Gothic UI", bold=True, size=14)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells("D3:F3")
    c = ws["D3"]; c.value = "株式会社 ライフアドバンス"
    c.font = Font(name="Yu Gothic UI", bold=True, size=11, color="1F3864")
    c.alignment = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[3].height = 26

    # ── Row4: キャッチコピー + 会社住所
    ws.merge_cells("A4:C4")
    c = ws["A4"]; c.value = "次の入居者獲得に向けた、付加価値向上のご提案です。"
    c.font = Font(name="Yu Gothic UI", size=10, color="2E7D32", bold=True)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells("D4:F4")
    c = ws["D4"]; c.value = "所在地：東京都渋谷区東3-25-11 TOKYU REIT恵比寿ビル4階"
    c.font = Font(name="Yu Gothic UI", size=8, color="444444")
    c.alignment = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[4].height = 20

    # ── Row5-7: 物件情報
    prop_rows = [("物件名", property_name), ("物件住所", property_address), ("担当", staff_name)]
    for idx, (label, val) in enumerate(prop_rows):
        rn = 5 + idx
        ws.merge_cells(start_row=rn, start_column=1, end_row=rn, end_column=2)
        c = ws.cell(row=rn, column=1, value=label)
        c.font = Font(name="Yu Gothic UI", size=9, bold=True)
        c.fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = make_border()
        ws.merge_cells(start_row=rn, start_column=3, end_row=rn, end_column=6)
        c = ws.cell(row=rn, column=3, value=val)
        c.font = Font(name="Yu Gothic UI", size=9)
        c.border = make_border()
        ws.row_dimensions[rn].height = 16

    ws.row_dimensions[8].height = 6  # スペーサー

    # ── Row9: 概要説明
    ws.merge_cells("A9:F9")
    c = ws["A9"]
    c.value = "【ご提案の目的】原状回復工事と同時に行うことで、工事費を抑えながら物件の魅力を高め、空室期間の短縮・賃料維持・次期入居者の獲得に繋げます。"
    c.font = Font(name="Yu Gothic UI", size=9, color="333333")
    c.fill = PatternFill(start_color="F1F8E9", end_color="F1F8E9", fill_type="solid")
    c.border = make_border()
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[9].height = 28

    ws.row_dimensions[10].height = 6

    # ── Row11-12: テーブルヘッダー
    row_h = 11
    ws.merge_cells(start_row=row_h, start_column=1, end_row=row_h+1, end_column=1)
    c = ws.cell(row=row_h, column=1, value="No."); apply_header_style(c, "2E7D32")
    ws.merge_cells(start_row=row_h, start_column=2, end_row=row_h+1, end_column=2)
    c = ws.cell(row=row_h, column=2, value="提案内容"); apply_header_style(c, "2E7D32")
    ws.merge_cells(start_row=row_h, start_column=3, end_row=row_h+1, end_column=3)
    c = ws.cell(row=row_h, column=3, value="工事内容・仕様"); apply_header_style(c, "2E7D32")
    ws.merge_cells(start_row=row_h, start_column=4, end_row=row_h+1, end_column=4)
    c = ws.cell(row=row_h, column=4, value="期待効果"); apply_header_style(c, "2E7D32")
    ws.merge_cells(start_row=row_h, start_column=5, end_row=row_h+1, end_column=5)
    c = ws.cell(row=row_h, column=5, value="概算費用（税込）"); apply_header_style(c, "2E7D32")
    ws.merge_cells(start_row=row_h, start_column=6, end_row=row_h+1, end_column=6)
    c = ws.cell(row=row_h, column=6, value="写真・参考イメージ"); apply_header_style(c, "2E7D32")
    ws.row_dimensions[row_h].height = 20
    ws.row_dimensions[row_h+1].height = 16

    # ── データ行
    PHOTO_ROW_HEIGHT = 90
    r = 13
    total_cost = 0
    temp_files = []

    for i, prop in enumerate(proposals):
        bg = "F9FBF0" if i % 2 == 0 else "FFFFFF"
        title = prop.get("title", "")
        work  = prop.get("work_content", "")
        effect = prop.get("expected_effect", "")
        cost  = int(float(prop.get("estimated_cost", 0) or 0))
        total_cost += cost

        for col in range(1, 7):
            c = ws.cell(row=r, column=col)
            c.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
            c.border = make_border()
            c.font = Font(name="Yu Gothic UI", size=10)

        ws.cell(row=r, column=1, value=i+1).alignment = Alignment(horizontal="center", vertical="top")
        c = ws.cell(row=r, column=2, value=title)
        c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        c.font = Font(name="Yu Gothic UI", size=10, bold=True)
        ws.cell(row=r, column=3, value=work).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws.cell(row=r, column=4, value=effect).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        c = ws.cell(row=r, column=5, value=cost)
        c.number_format = '#,##0"円"'; c.alignment = Alignment(horizontal="right", vertical="top")

        # 写真埋め込み
        photo_bytes = photos.get(i)
        if photo_bytes:
            try:
                with tempfile.NamedTemporaryFile(suffix=".jpg", delete=False) as tmp:
                    # JPEG変換
                    img = PILImage.open(BytesIO(photo_bytes))
                    img = img.convert("RGB")
                    # 縮小（最大 220x160px）
                    img.thumbnail((220, 160), PILImage.LANCZOS)
                    img.save(tmp.name, "JPEG", quality=85)
                    temp_files.append(tmp.name)
                xl_img = XLImage(tmp.name)
                xl_img.width = 180
                xl_img.height = 130
                col_letter = get_column_letter(6)
                ws.add_image(xl_img, f"{col_letter}{r}")
                ws.row_dimensions[r].height = PHOTO_ROW_HEIGHT
            except Exception:
                ws.cell(row=r, column=6, value="※写真を手動で挿入してください")
                ws.row_dimensions[r].height = 50
        else:
            ws.cell(row=r, column=6, value="※写真を手動で挿入してください")
            ws.cell(row=r, column=6).font = Font(name="Yu Gothic UI", size=8, color="999999")
            ws.row_dimensions[r].height = 50

        r += 1

    # ── 合計行
    fill_row_borders(ws, r, 1, 6, "E8F5E9")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    c = ws.cell(row=r, column=1, value="提案合計（税込）")
    c.font = Font(name="Yu Gothic UI", bold=True, size=11)
    c.fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    c.alignment = Alignment(horizontal="right", vertical="center"); c.border = make_border()
    c = ws.cell(row=r, column=5, value=total_cost)
    c.number_format = '#,##0"円"'
    c.font = Font(name="Yu Gothic UI", bold=True, size=12, color="2E7D32")
    apply_data_style(c, bg_color="E8F5E9", bold=True, align="right")
    ws.row_dimensions[r].height = 22
    r += 2

    # ── 備考・クロージング
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    c = ws.cell(row=r, column=1, value="【ご注意】上記概算費用は現時点での目安です。現地確認後に正式なお見積りをご提出いたします。")
    c.font = Font(name="Yu Gothic UI", size=9, color="666666")
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[r].height = 18
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    c = ws.cell(row=r, column=1,
                value="ご不明な点・ご要望はお気軽にご相談ください。恵比寿不動産　TEL: 03-6421-0544")
    c.font = Font(name="Yu Gothic UI", size=9, color="1F3864", bold=True)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[r].height = 18

    ws.freeze_panes = "A13"

    # 一時ファイル削除
    for tmp_path in temp_files:
        try:
            os.unlink(tmp_path)
        except Exception:
            pass


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5001))
    app.run(host="0.0.0.0", port=port, use_reloader=False)
